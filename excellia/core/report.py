"""Reports: highlighted xlsx export + the Data Health Score.

Ports what the legacy GUI's export did, engine-side. Highlight priority
when a cell has several problems (the legacy order): outlier >
duplicate > mixed > format > missing. The health score is the legacy
heuristic — start at 100, weighted deduction per percent of affected
rows — always returned WITH its breakdown, never a bare number.
"""

from __future__ import annotations

import os
import time
from typing import Any

import pandas as pd

from excellia.core import anomaly, ingest, validate
from excellia.core.models import Flag, Issue

# category -> (deduction weight per percent of rows affected, fill colour)
_CATEGORIES: dict[str, tuple[float, str]] = {
    "outlier":   (0.8, "FFC7CE"),  # red
    "duplicate": (0.2, "FFD966"),  # amber
    "mixed":     (0.5, "BDD7EE"),  # blue
    "format":    (0.3, "F8CBAD"),  # orange
    "missing":   (0.4, "E7E6E6"),  # grey
}
_PRIORITY = ("outlier", "duplicate", "mixed", "format", "missing")

_MAX_DATA_ROWS = 100_000  # keep the Data sheet openable; issues sheets are complete


def _category(name: str) -> str | None:
    """Highlight category for a rule_name (Issue) or kind (Flag)."""
    if name in ("multivariate_outlier", "column_outlier", "pattern_break"):
        return "outlier"
    if name in ("duplicate_row", "duplicate_id", "unique", "near_duplicate"):
        return "duplicate"
    if name == "mixed_types":
        return "mixed"
    if name.startswith("format_") or name == "reference":
        return "format"
    if name in ("missing_value", "required_field"):
        return "missing"
    return None


def health_score(row_count: int, issues: list[Issue], flags: list[Flag]) -> dict[str, Any]:
    """The legacy heuristic with its working shown.

    100 minus (weight x percent of rows affected) per category:
    outliers 0.8/pct, mixed 0.5/pct, missing 0.4/pct, format 0.3/pct,
    duplicates 0.2/pct."""
    rows_by_cat: dict[str, set[int]] = {c: set() for c in _CATEGORIES}
    for issue in issues:
        cat = _category(issue.rule_name)
        if cat:
            rows_by_cat[cat].add(issue.row)
    for flag in flags:
        cat = _category(flag.kind)
        if cat:
            rows_by_cat[cat].add(flag.row)

    breakdown = []
    total_deduction = 0.0
    for cat in _PRIORITY:
        affected = len(rows_by_cat[cat])
        if not affected or not row_count:
            continue
        pct = 100.0 * affected / row_count
        weight = _CATEGORIES[cat][0]
        deduction = round(weight * pct, 2)
        total_deduction += deduction
        breakdown.append({
            "category": cat, "rows_affected": affected,
            "pct_of_rows": round(pct, 2), "weight_per_pct": weight,
            "deduction": deduction,
        })
    return {
        "score": max(0, round(100 - total_deduction)),
        "breakdown": breakdown,
        "note": "score = 100 - sum(weight x pct rows affected) per category",
    }


def _cell_categories(issues: list[Issue], flags: list[Flag]) -> dict[tuple[int, str], str]:
    """(excel_row, column) -> winning highlight category. Column '*' means whole row."""
    chosen: dict[tuple[int, str], str] = {}

    def offer(row: int, col: str, cat: str | None) -> None:
        if cat is None:
            return
        key = (row, col)
        if key not in chosen or _PRIORITY.index(cat) < _PRIORITY.index(chosen[key]):
            chosen[key] = cat

    for issue in issues:
        offer(issue.row, issue.column, _category(issue.rule_name))
    for flag in flags:
        cat = _category(flag.kind)
        if flag.columns:
            for col in flag.columns:
                offer(flag.row, col, cat)
        else:
            offer(flag.row, "*", cat)
    return chosen


def export_report(
    file_path: str,
    out_path: str | None = None,
    ruleset: str | dict = "default",
    sensitivity: float = 0.05,
    sheet: str | None = None,
) -> dict[str, Any]:
    """Run validate + anomalies and write a highlighted xlsx report.

    Sheets: Data (cells coloured by issue category), Issues, Anomalies,
    Summary (health score with breakdown, counts, legend). Always writes
    a NEW file — never touches the input. Returns {path, health, counts}.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    df = ingest.load(file_path, sheet=sheet)
    issues = validate.validate(df, ruleset=ruleset)
    flags = anomaly.detect_anomalies(df, sensitivity=sensitivity)
    health = health_score(len(df), issues, flags)

    if out_path is None:
        stem, _ = os.path.splitext(file_path)
        out_path = f"{stem}_excellia_report.xlsx"
    if os.path.abspath(out_path) == os.path.abspath(file_path):
        raise ValueError("Refusing to overwrite the input file; pick another out_path")
    if os.path.exists(out_path):
        stem, ext = os.path.splitext(out_path)
        out_path = f"{stem}_{int(time.time())}{ext}"

    wb = Workbook()
    bold = Font(bold=True)
    fills = {cat: PatternFill("solid", fgColor=color)
             for cat, (_, color) in _CATEGORIES.items()}

    # --- Data sheet with highlights ---
    ws = wb.active
    ws.title = "Data"
    ws.append(list(df.columns))
    for cell in ws[1]:
        cell.font = bold
    col_index = {c: i + 1 for i, c in enumerate(df.columns)}
    chosen = _cell_categories(issues, flags)
    shown = df.head(_MAX_DATA_ROWS)
    for record in shown.itertuples(index=False):
        ws.append([None if pd.isna(v) else v for v in record])
    for (row, col), cat in chosen.items():
        if row - 1 > len(shown):  # beyond the cap (row is Excel-numbered)
            continue
        if col == "*":
            for c in range(1, len(df.columns) + 1):
                ws.cell(row=row, column=c).fill = fills[cat]
        elif col in col_index:
            ws.cell(row=row, column=col_index[col]).fill = fills[cat]
    if len(df) > _MAX_DATA_ROWS:
        ws.append([f"... truncated: showing first {_MAX_DATA_ROWS} of {len(df)} rows "
                   "(Issues/Anomalies sheets are complete)"])

    # --- Issues sheet ---
    ws = wb.create_sheet("Issues")
    ws.append(["row", "column", "rule", "severity", "reason", "value"])
    for cell in ws[1]:
        cell.font = bold
    for i in issues:
        ws.append([i.row, i.column, i.rule_name, i.severity, i.reason,
                   None if i.value is None else str(i.value)])

    # --- Anomalies sheet ---
    ws = wb.create_sheet("Anomalies")
    ws.append(["row", "kind", "confidence", "columns", "reason"])
    for cell in ws[1]:
        cell.font = bold
    for f in flags:
        ws.append([f.row, f.kind, f.confidence, ", ".join(f.columns), f.reason])

    # --- Summary sheet ---
    ws = wb.create_sheet("Summary")
    ws.append(["Data Health Score", health["score"]])
    ws["A1"].font = ws["B1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["category", "rows affected", "% of rows", "weight/pct", "deduction"])
    for cell in ws[3]:
        cell.font = bold
    for item in health["breakdown"]:
        ws.append([item["category"], item["rows_affected"], item["pct_of_rows"],
                   item["weight_per_pct"], item["deduction"]])
    ws.append([])
    ws.append(["rows", len(df)])
    ws.append(["issues (rule violations)", len(issues)])
    ws.append(["anomaly flags", len(flags)])
    ws.append([])
    ws.append(["highlight legend"])
    ws.cell(row=ws.max_row, column=1).font = bold
    for cat in _PRIORITY:
        ws.append([cat])
        ws.cell(row=ws.max_row, column=1).fill = fills[cat]

    wb.save(out_path)
    return {
        "path": os.path.abspath(out_path),
        "health": health,
        "issues": len(issues),
        "flags": len(flags),
        "note": "row numbers are Excel rows: header is row 1, first data row is 2",
    }


def reconciliation_report(result, summary: dict, out_path: str) -> str:
    """The 5-sheet reconciliation xlsx: Summary / Matched / Only-in-A /
    Only-in-B / Discrepancies (side-by-side a|b with variance)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    bold = Font(bold=True)

    ws = wb.active
    ws.title = "Summary"
    rows = [
        ("matched", summary.get("matched")),
        ("discrepancies", summary.get("discrepancies")),
        ("only_in_a", summary.get("only_in_a")),
        ("only_in_b", summary.get("only_in_b")),
        ("match_rate", summary.get("match_rate")),
        ("profile", summary.get("profile")),
    ]
    for label, value in rows:
        ws.append([label, value])
    ws.append([])
    ws.append(["match levels"])
    ws.cell(row=ws.max_row, column=1).font = bold
    for level, n in sorted((summary.get("match_levels") or {}).items()):
        ws.append([level, n])
    ws.append([])
    ws.append(["variance totals (abs)"])
    ws.cell(row=ws.max_row, column=1).font = bold
    for col, total in sorted((summary.get("variance_totals_abs") or {}).items()):
        ws.append([col, total])

    def _sheet_of_records(title: str, records: list[dict]) -> None:
        ws = wb.create_sheet(title)
        if not records:
            ws.append(["(empty)"])
            return
        headers = list(records[0])
        ws.append(headers)
        for cell in ws[1]:
            cell.font = bold
        for rec in records:
            ws.append([_plain(rec.get(h)) for h in headers])

    _sheet_of_records("Matched", result.matched)
    _sheet_of_records("Only-in-A", result.only_in_a)
    _sheet_of_records("Only-in-B", result.only_in_b)

    ws = wb.create_sheet("Discrepancies")
    if result.discrepancies:
        keys = [k for k in result.discrepancies[0]
                if k not in ("a", "b", "differences")]
        ws.append(keys + ["field", "a", "b", "diff_abs", "diff_pct"])
        for cell in ws[1]:
            cell.font = bold
        for d in result.discrepancies:
            for col, diff in d.get("differences", {}).items():
                ws.append([_plain(d.get(k)) for k in keys] + [
                    col, _plain(diff.get("a")), _plain(diff.get("b")),
                    diff.get("diff_abs"), diff.get("diff_pct")])
    else:
        ws.append(["(empty)"])

    wb.save(out_path)
    return os.path.abspath(out_path)


def _plain(value):
    """Cell-safe scalar (dicts/lists become compact JSON strings)."""
    if isinstance(value, (dict, list)):
        import json

        return json.dumps(value, default=str)
    return value
