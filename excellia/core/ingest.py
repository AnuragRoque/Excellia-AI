"""Ingest & profiling: load a spreadsheet and describe its shape.

Lifted from the old Excellia GUI's ``load_dataframe`` (routes.py) and
generalised: no Flask, no session state, no print-based logging.
"""

from __future__ import annotations

import os
from collections import Counter
from typing import Iterator

import pandas as pd

from excellia.core.models import ColumnProfile, Profile
from excellia.core.rules.builtin import FORMATS

SUPPORTED_EXTENSIONS = (".xlsx", ".xlsm", ".xls", ".csv", ".tsv")

_CSV_ENCODINGS = ("utf-8", "utf-8-sig", "cp1252", "latin1")
_CSV_DELIMITERS = (",", ";", "\t", "|")

# Values treated as empty besides real NaN (mirrors the GUI's is_empty_value)
_EMPTY_STRINGS = {"", "nan", "null", "none", "n/a", "na"}


class IngestError(ValueError):
    """Raised when a file cannot be loaded as tabular data."""


def _read_csv(file_path: str, delimiter: str | None) -> pd.DataFrame | None:
    """One robust CSV read attempt per encoding; None if all fail."""
    for encoding in _CSV_ENCODINGS:
        try:
            df = pd.read_csv(
                file_path,
                encoding=encoding,
                sep=delimiter,
                engine="python",
                on_bad_lines="skip",
            )
        except Exception:
            continue
        # Reject fake parses: a single wide column whose name still
        # contains a delimiter means we picked the wrong separator.
        if df.shape[1] > 1 or (
            df.shape[1] == 1
            and not any(d in str(df.columns[0]) for d in _CSV_DELIMITERS)
        ):
            return df
    return None


def load(file_path: str, sheet: str | None = None) -> pd.DataFrame:
    """Load a spreadsheet into a DataFrame.

    Auto-detects delimiter and encoding for CSV/TSV. Raises
    FileNotFoundError for missing files and IngestError for
    unsupported or unparseable files.
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(file_path)
    ext = os.path.splitext(file_path)[1].lower()
    if ext not in SUPPORTED_EXTENSIONS:
        raise IngestError(
            f"Unsupported file type '{ext}'. Supported: {', '.join(SUPPORTED_EXTENSIONS)}"
        )

    if ext in (".xlsx", ".xlsm", ".xls"):
        engine = "xlrd" if ext == ".xls" else "openpyxl"
        df = pd.read_excel(file_path, sheet_name=sheet or 0, engine=engine)
    elif ext == ".tsv":
        df = _read_csv(file_path, "\t")
    else:  # .csv — sniff first, then try explicit delimiters
        df = _read_csv(file_path, None)
        if df is None or df.empty:
            for delimiter in _CSV_DELIMITERS:
                df = _read_csv(file_path, delimiter)
                if df is not None and not df.empty:
                    break

    if df is None or df.empty:
        raise IngestError(f"Could not parse any tabular data from {file_path}")

    df.columns = [str(c) for c in df.columns]
    return df


def nonempty(series: pd.Series) -> pd.Series:
    """Values that are neither NaN nor an empty-like string."""
    s = series.dropna()
    if s.empty:
        return s
    mask = ~s.astype(str).str.strip().str.lower().isin(_EMPTY_STRINGS)
    return s[mask]


def _detect_format(values: pd.Series) -> str | None:
    """Name of the built-in format that >=90% of values match, if any."""
    sample = values.astype(str).str.strip().head(500)
    if sample.empty:
        return None
    for name, pattern in FORMATS.items():
        hits = sample.apply(lambda v: bool(pattern.fullmatch(v))).mean()
        if hits >= 0.9:
            return name
    return None


def _infer_type(
    series: pd.Series, values: pd.Series, detected_format: str | None
) -> str:
    if detected_format in ("gst", "pan", "aadhaar", "ifsc"):
        return "id"
    if pd.api.types.is_numeric_dtype(series):
        return "number"
    if pd.api.types.is_datetime64_any_dtype(series):
        return "date"
    if values.empty:
        return "text"

    sample = values.astype(str).str.strip().head(500)
    # currency: numbers wearing a symbol
    if sample.str.match(r"^[₹$€£]\s*-?[\d,]+(\.\d+)?$").mean() >= 0.8:
        return "currency"
    if pd.to_numeric(sample.str.replace(",", ""), errors="coerce").notna().mean() >= 0.9:
        return "number"
    try:
        if pd.to_datetime(sample, errors="coerce", format="mixed").notna().mean() >= 0.9:
            return "date"
    except (ValueError, TypeError):
        pass

    unique_ratio = values.nunique() / len(values)
    if unique_ratio > 0.95 and len(values) > 10:
        return "id"
    if values.nunique() <= max(20, int(0.05 * len(values))):
        return "categorical"
    return "text"


def profile(file_path: str, sheet: str | None = None) -> Profile:
    """Profile a spreadsheet: row/col counts, per-column type inference,
    null rates, cardinality, min/max/mean, top values, detected formats."""
    df = load(file_path, sheet=sheet)
    columns: list[ColumnProfile] = []

    for col in df.columns:
        series = df[col]
        values = nonempty(series)
        null_rate = 1.0 - (len(values) / len(series)) if len(series) else 0.0
        detected_format = _detect_format(values) if not values.empty else None
        inferred = _infer_type(series, values, detected_format)

        col_min = col_max = None
        col_mean = None
        numeric = pd.to_numeric(values, errors="coerce").dropna()
        if not numeric.empty:
            col_min, col_max = numeric.min().item(), numeric.max().item()
            col_mean = round(float(numeric.mean()), 4)
        elif not values.empty:
            try:
                col_min, col_max = str(values.min()), str(values.max())
            except TypeError:
                pass  # mixed types that won't order

        top = values.astype(str).value_counts().head(5).index.tolist()
        columns.append(
            ColumnProfile(
                name=col,
                inferred_type=inferred,
                null_rate=round(null_rate, 4),
                cardinality=int(values.nunique()),
                min=col_min,
                max=col_max,
                mean=col_mean,
                top_values=top,
                detected_format=detected_format,
            )
        )

    return Profile(
        file=file_path,
        sheet=sheet,
        row_count=len(df),
        column_count=len(df.columns),
        columns=columns,
    )


# --- big files: chunked streaming ------------------------------------
#
# The in-memory path above is the source of truth for small files.
# iter_chunks/profile_large exist so a 500K-row file never has to fit
# in RAM at once: xlsx streams via openpyxl read-only mode, csv via
# pandas chunked readers. Legacy .xls has no streaming reader, so it
# falls back to a full load sliced into chunks (those files are small
# by format era anyway).

DEFAULT_CHUNK_SIZE = 50_000


def _sniff_csv(file_path: str) -> tuple[str, str]:
    """(encoding, delimiter) that best parses the header line."""
    for encoding in _CSV_ENCODINGS:
        try:
            with open(file_path, encoding=encoding) as f:
                header = f.readline()
        except (UnicodeDecodeError, OSError):
            continue
        counts = {d: header.count(d) for d in _CSV_DELIMITERS}
        best = max(counts, key=counts.get)  # type: ignore[arg-type]
        return encoding, best if counts[best] else ","
    raise IngestError(f"Could not decode {file_path} with any known encoding")


def iter_chunks(
    file_path: str, sheet: str | None = None, chunk_size: int = DEFAULT_CHUNK_SIZE
) -> Iterator[pd.DataFrame]:
    """Yield a spreadsheet as DataFrame chunks without loading it whole.

    Every chunk shares the same string column names. Raises the same
    instructive errors as ``load``.
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(file_path)
    ext = os.path.splitext(file_path)[1].lower()
    if ext not in SUPPORTED_EXTENSIONS:
        raise IngestError(
            f"Unsupported file type '{ext}'. Supported: {', '.join(SUPPORTED_EXTENSIONS)}"
        )

    if ext in (".xlsx", ".xlsm"):
        import openpyxl

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        try:
            ws = wb[sheet] if sheet else wb.active
        except KeyError:
            wb.close()
            raise IngestError(
                f"Worksheet '{sheet}' not found. Sheets: {', '.join(wb.sheetnames)}"
            )
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            wb.close()
            raise IngestError(f"Could not parse any tabular data from {file_path}")
        columns = [str(c) for c in header]
        batch: list[tuple] = []
        try:
            for row in rows:
                batch.append(row)
                if len(batch) >= chunk_size:
                    yield pd.DataFrame(batch, columns=columns)
                    batch = []
            if batch:
                yield pd.DataFrame(batch, columns=columns)
        finally:
            wb.close()
    elif ext in (".csv", ".tsv"):
        encoding, delimiter = ("utf-8", "\t") if ext == ".tsv" else _sniff_csv(file_path)
        reader = pd.read_csv(
            file_path, encoding=encoding, sep=delimiter, chunksize=chunk_size,
            on_bad_lines="skip",
        )
        for chunk in reader:
            chunk.columns = [str(c) for c in chunk.columns]
            yield chunk
    else:  # .xls — no streaming reader exists; load once, slice
        df = load(file_path, sheet=sheet)
        for start in range(0, len(df), chunk_size):
            yield df.iloc[start : start + chunk_size]


class _ColumnAccumulator:
    """Streaming per-column stats merged across chunks."""

    __slots__ = ("n", "n_nonempty", "sample", "counter", "uniques", "num_min",
                 "num_max", "num_sum", "num_count")

    def __init__(self) -> None:
        self.n = 0                     # total cells incl. empty
        self.n_nonempty = 0
        self.sample: list[str] = []    # first nonempty values, for type/format inference
        self.counter: Counter = Counter()
        self.uniques: set[str] = set()
        self.num_min = self.num_max = None
        self.num_sum = 0.0
        self.num_count = 0

    def update(self, series: pd.Series) -> None:
        self.n += len(series)
        values = nonempty(series)
        if values.empty:
            return
        self.n_nonempty += len(values)
        as_str = values.astype(str).str.strip()
        if len(self.sample) < 1000:
            self.sample.extend(as_str.head(1000 - len(self.sample)).tolist())
        self.counter.update(as_str.tolist())
        if len(self.counter) > 50_000:  # cap memory; keeps top values approximate
            self.counter = Counter(dict(self.counter.most_common(25_000)))
        self.uniques.update(as_str.tolist())
        numeric = pd.to_numeric(values, errors="coerce").dropna()
        if not numeric.empty:
            lo, hi = float(numeric.min()), float(numeric.max())
            self.num_min = lo if self.num_min is None else min(self.num_min, lo)
            self.num_max = hi if self.num_max is None else max(self.num_max, hi)
            self.num_sum += float(numeric.sum())
            self.num_count += len(numeric)


def profile_large(
    file_path: str, sheet: str | None = None, chunk_size: int = DEFAULT_CHUNK_SIZE
) -> Profile:
    """Streaming ``profile`` for files too big to hold in memory.

    Same output shape as ``profile``. Honest approximations: type and
    format inference use the first ~1000 nonempty values per column,
    and top_values become approximate past 50K distinct values.
    """
    accs: dict[str, _ColumnAccumulator] = {}
    order: list[str] = []
    row_count = 0
    for chunk in iter_chunks(file_path, sheet=sheet, chunk_size=chunk_size):
        row_count += len(chunk)
        for col in chunk.columns:
            if col not in accs:
                accs[col] = _ColumnAccumulator()
                order.append(col)
            accs[col].update(chunk[col])
    if row_count == 0:
        raise IngestError(f"Could not parse any tabular data from {file_path}")

    columns: list[ColumnProfile] = []
    for col in order:
        acc = accs[col]
        sample = pd.Series(acc.sample, dtype=object)
        detected_format = _detect_format(sample) if not sample.empty else None
        inferred = _infer_type(sample, sample, detected_format)
        col_min: object = acc.num_min
        col_max: object = acc.num_max
        col_mean = (
            round(acc.num_sum / acc.num_count, 4) if acc.num_count else None
        )
        if col_min is None and acc.uniques:
            try:
                col_min, col_max = min(acc.uniques), max(acc.uniques)
            except TypeError:
                pass
        n_nonempty = acc.n_nonempty
        columns.append(
            ColumnProfile(
                name=col,
                inferred_type=inferred,
                null_rate=round(1.0 - (n_nonempty / acc.n), 4) if acc.n else 0.0,
                cardinality=len(acc.uniques),
                min=col_min,
                max=col_max,
                mean=col_mean,
                top_values=[v for v, _ in acc.counter.most_common(5)],
                detected_format=detected_format,
            )
        )
    return Profile(
        file=file_path, sheet=sheet, row_count=row_count,
        column_count=len(order), columns=columns,
    )
