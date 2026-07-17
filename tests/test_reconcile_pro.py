"""Reconciliation pro: match levels, variance, fuzzy keys, profiles,
and the 5-sheet report — CMS-vs-Switch style fixtures."""

import pandas as pd
import pytest
from openpyxl import load_workbook

from excellia.core import reconcile, report, store


@pytest.fixture()
def cms():
    return pd.DataFrame({
        "txn_id": ["T001", "T002", "T003", "T004", "T005"],
        "amount": [100.00, 250.00, 75.50, 900.00, 40.00],
        "merchant": ["Acme Stores", "Zen Corp", "Iota Ltd", "Big Bazaar", "Chai Point"],
    })


@pytest.fixture()
def switch():
    return pd.DataFrame({
        "txn_id": ["T001", "T002", "T0O3", "T004", "T999"],  # T0O3: typo'd key (O for 0)
        "amount": [100.00, 250.02, 75.50, 850.00, 60.00],
        "merchant": ["Acme Stores", "Zen Corp", "Iota Ltd", "Big Bazaar", "Ghost Shop"],
    })


def test_match_levels(cms, switch):
    result = reconcile.reconcile(cms, switch, keys=["txn_id"],
                                 tolerance={"numeric": 0.05})
    levels = {r["txn_id"]: r["match_level"] for r in result.matched}
    assert levels["T001"] == "L1"          # exact everywhere
    assert levels["T002"] == "L2"          # amount differs by 0.02, inside tolerance
    disc = {d["txn_id"]: d for d in result.discrepancies}
    assert "T004" in disc                  # amount differs by 50, outside tolerance


def test_variance_columns(cms, switch):
    result = reconcile.reconcile(cms, switch, keys=["txn_id"])
    disc = {d["txn_id"]: d for d in result.discrepancies}
    diff = disc["T004"]["differences"]["amount"]
    assert diff["diff_abs"] == -50.0
    assert diff["diff_pct"] == pytest.approx(-5.5556, abs=0.001)


def test_fuzzy_keys_l3(cms, switch):
    strict = reconcile.reconcile(cms, switch, keys=["txn_id"])
    assert any(r["txn_id"] == "T003" for r in strict.only_in_a)

    fuzzy = reconcile.reconcile(cms, switch, keys=["txn_id"], fuzzy_keys=0.7)
    l3 = [r for r in fuzzy.matched if r.get("match_level") == "L3"]
    assert len(l3) == 1 and l3[0]["txn_id"] == "T003"
    assert l3[0]["key_similarity"] >= 0.7
    assert not any(r["txn_id"] == "T003" for r in fuzzy.only_in_a)
    # the truly unmatched stay unmatched
    assert any(r["txn_id"] == "T999" for r in fuzzy.only_in_b)


def test_run_profile_with_presteps_and_dedupe(cms, switch):
    # duplicate txn in the switch feed that must be aggregated before matching
    switch2 = pd.concat([switch, pd.DataFrame({
        "txn_id": ["T001"], "amount": [0.00], "merchant": ["Acme Stores"]})],
        ignore_index=True)
    profile = {
        "name": "monthly-cms-switch",
        "keys": ["txn_id"],
        "tolerance": {"numeric": 0.05},
        "pre_recipe_b": [{"op": "trim", "params": {"columns": ["merchant"]}}],
        "dedupe_b": {"columns": ["txn_id"], "aggregate": {"amount": "sum"}},
    }
    run = reconcile.run_profile(cms, switch2, profile)
    summary = run["summary"]
    assert summary["profile"] == "monthly-cms-switch"
    assert summary["match_rate"] is not None
    assert "L1" in summary["match_levels"] or "L2" in summary["match_levels"]
    assert "amount" in summary["variance_totals_abs"]


def test_run_profile_needs_keys(cms, switch):
    with pytest.raises(ValueError, match="keys"):
        reconcile.run_profile(cms, switch, {"tolerance": {}})


def test_reconciliation_report_five_sheets(cms, switch, tmp_path):
    run = reconcile.run_profile(cms, switch, {"keys": ["txn_id"],
                                              "tolerance": {"numeric": 0.05},
                                              "name": "t"})
    out = str(tmp_path / "recon.xlsx")
    path = report.reconciliation_report(run["result"], run["summary"], out)
    wb = load_workbook(path)
    assert wb.sheetnames == ["Summary", "Matched", "Only-in-A", "Only-in-B",
                             "Discrepancies"]
    matched = wb["Matched"]
    headers = [c.value for c in matched[1]]
    assert "match_level" in headers
    disc = wb["Discrepancies"]
    headers = [c.value for c in disc[1]]
    assert {"field", "a", "b", "diff_abs", "diff_pct"} <= set(headers)
    summary_labels = [row[0].value for row in wb["Summary"].iter_rows(max_col=1)]
    assert "match_rate" in summary_labels


def test_profile_store_roundtrip():
    spec = {"keys": ["txn_id"], "tolerance": {"numeric": 0.01}, "name": "p1"}
    store.save("profiles", "p1", spec)
    assert store.load("profiles", "p1") == spec
    assert "p1" in store.list_names("profiles")
