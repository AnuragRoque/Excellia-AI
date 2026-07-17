"""report.py: the xlsx opens, the highlights land, the score shows its work."""

import os

import pytest
from openpyxl import load_workbook

from excellia.core import report
from excellia.core.models import Flag, Issue

EXAMPLES = os.path.join(os.path.dirname(__file__), "..", "examples")
MESSY = os.path.join(EXAMPLES, "messy_vendors.xlsx")


def test_health_score_math():
    # 100 rows: 10 with outliers (0.8*10=8), 5 missing (0.4*5=2) -> 90
    issues = [Issue(r, "a", "missing_value", "warning", "x") for r in range(2, 7)]
    flags = [Flag(r, "column_outlier", 0.9, "x") for r in range(10, 20)]
    health = report.health_score(100, issues, flags)
    assert health["score"] == 90
    by_cat = {b["category"]: b for b in health["breakdown"]}
    assert by_cat["outlier"]["deduction"] == 8.0
    assert by_cat["missing"]["deduction"] == 2.0
    assert "note" in health


def test_health_score_priority_and_floor():
    # every row broken badly -> score clamps at 0, never negative
    flags = [Flag(r, "multivariate_outlier", 1.0, "x") for r in range(2, 102)]
    issues = [Issue(r, "a", "mixed_types", "warning", "x") for r in range(2, 102)]
    assert report.health_score(100, issues, flags)["score"] == 0


def test_category_mapping():
    assert report._category("format_gst") == "format"
    assert report._category("duplicate_row") == "duplicate"
    assert report._category("near_duplicate") == "duplicate"
    assert report._category("required_field") == "missing"
    assert report._category("pattern_break") == "outlier"
    assert report._category("something_new") is None


def test_export_report_roundtrip(tmp_path):
    out = str(tmp_path / "report.xlsx")
    result = report.export_report(MESSY, out_path=out)
    assert result["path"] == os.path.abspath(out)
    assert 0 <= result["health"]["score"] <= 100
    assert result["issues"] > 0 and result["flags"] > 0

    wb = load_workbook(out)
    assert wb.sheetnames == ["Data", "Issues", "Anomalies", "Summary"]
    data = wb["Data"]
    assert data.max_row >= 51  # header + 50 data rows
    filled = sum(
        1 for row in data.iter_rows()
        for cell in row if cell.fill and cell.fill.fill_type == "solid"
    )
    assert filled > 0, "no highlighted cells in the Data sheet"
    issues_ws = wb["Issues"]
    assert issues_ws.max_row - 1 == result["issues"]
    summary = wb["Summary"]
    assert summary["A1"].value == "Data Health Score"
    assert summary["B1"].value == result["health"]["score"]


def test_export_refuses_overwriting_input(tmp_path):
    import shutil

    target = tmp_path / "copy.xlsx"
    shutil.copy(MESSY, target)
    with pytest.raises(ValueError, match="Refusing to overwrite"):
        report.export_report(str(target), out_path=str(target))


def test_export_never_clobbers_existing_report(tmp_path):
    out = str(tmp_path / "r.xlsx")
    first = report.export_report(MESSY, out_path=out)
    second = report.export_report(MESSY, out_path=out)
    assert first["path"] != second["path"]
    assert os.path.exists(first["path"]) and os.path.exists(second["path"])
